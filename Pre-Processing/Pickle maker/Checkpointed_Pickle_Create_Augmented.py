import pickle
import gzip
import os
import shutil
import tempfile
import re
import cv2
import numpy as np
import torch
from concurrent.futures import ProcessPoolExecutor, as_completed
from tensorflow.keras.applications import EfficientNetB7
from tensorflow.keras.layers import GlobalAveragePooling2D
from tensorflow.keras.models import Model
from tensorflow.keras.applications.efficientnet import preprocess_input
from tensorflow.keras.preprocessing.image import ImageDataGenerator
from openpyxl import load_workbook
import portalocker
import datetime
import multiprocessing
import logging
import time

# Initialize logging
logging.basicConfig(filename='pickle_creation.log', level=logging.INFO, 
                    format='%(asctime)s:%(levelname)s:%(message)s')

def initialize_model():
    base_model = EfficientNetB7(weights='imagenet', include_top=False)
    x = base_model.output
    x = GlobalAveragePooling2D()(x)
    feature_extractor = Model(inputs=base_model.input, outputs=x)
    return feature_extractor

def save_checkpoint(checkpoint_path, checkpoint_name, list_of_inputs):
    unstable_path = os.path.join(checkpoint_path, "unstable")
    unstable_file = os.path.join(unstable_path, checkpoint_name)
    if not os.path.exists(unstable_path):
        os.makedirs(unstable_path)
    
    with tempfile.NamedTemporaryFile(delete=False) as tmp_file:
        with gzip.GzipFile(fileobj=tmp_file, mode='wb') as gz:
            pickle.dump(list_of_inputs, gz)
        tmp_filename = tmp_file.name

    try:
        shutil.move(tmp_filename, unstable_file)
        shutil.move(unstable_file, os.path.join(checkpoint_path, checkpoint_name))
        logging.info("Backup made at: " + str(checkpoint_path))

        timestamp = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
        backup_folder = os.path.join(checkpoint_path, "backup")
        if not os.path.exists(backup_folder):
            os.makedirs(backup_folder)
        backup_unstable_folder = os.path.join(backup_folder, f"unstable_{checkpoint_name}_{timestamp}")
        shutil.copytree(unstable_path, backup_unstable_folder)
        logging.info("Unstable folder backed up at: " + str(backup_unstable_folder))
    except Exception as e:
        logging.error("Error!!!! Could not create backup: " + str(e))
        exit()

def load_checkpoint(checkpoint_path, checkpoint_name):
    backup_file = os.path.join(checkpoint_path, checkpoint_name)
    retry_attempts = 60  # Number of retry attempts
    retry_delay = 40  # Delay in seconds (20 seconds)

    for attempt in range(retry_attempts):
        if os.path.exists(backup_file):
            logging.info("Loading from: " + str(backup_file))
            logging.info("\n*************************\n*************************\n*************************\n*** Checkpoint Loaded ***\n*************************\n*************************\n*************************\n")
            try:
                with open(backup_file, 'rb') as f:
                    portalocker.lock(f, portalocker.LOCK_SH)
                    try:
                        with gzip.GzipFile(fileobj=f) as gz:
                            data = pickle.load(gz)
                    finally:
                        portalocker.unlock(f)
                        logging.info("File unlocked successfully")
                if data:
                    last_entry = data.pop()  # Remove the last entry
                    logging.info(f"Reprocessing entry: {last_entry['name']}")
                else:
                    last_entry = None
                processed_names = {entry['name'] for entry in data}
                return data, processed_names
            except Exception as e:
                logging.error(f"Failed to read and lock the file: {e}")
                if attempt < retry_attempts - 1:
                    logging.info(f"Retrying in {retry_delay} seconds...")
                    time.sleep(retry_delay)
                else:
                    logging.error("Maximum retry attempts reached. Could not lock the file.")
                    raise e
        else:
            logging.info("Creating at: " + str(backup_file))
            logging.info("\n****************************************\n****************************************\n****************************************\n*** Checkpoint Loading Failed!!!!!!! ***\n****************************************\n****************************************\n****************************************\n")
            return None, set()

def get_features(filename, destination, feature_extractor, augment_params, num_augmentations):
    input_string = filename
    pattern = r'\d+'
    match = re.search(pattern, input_string)
    if match:
        first_match = match.group()
        input_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), destination, first_match, input_string)
        try:
            file_paths_frames = [file for file in sorted(os.listdir(input_folder)) if file.endswith(".jpg")]
        except:
            return None

        all_features = []
        datagen = ImageDataGenerator(
            brightness_range=[1 - augment_params['brightness'], 1 + augment_params['brightness']],
            rotation_range=augment_params['rotation'],
            width_shift_range=augment_params['width_shift'],
            height_shift_range=augment_params['height_shift'],
            zoom_range=augment_params['zoom'],
            shear_range=augment_params['shear']
        )

        for aug_index in range(num_augmentations):
            features_listofList = []
            for indx, frame_file in enumerate(file_paths_frames):
                frame_filename = os.path.join(input_folder, frame_file)
                image = cv2.imread(frame_filename)
                
                # Resize image to 600x600 using interpolation
                image = cv2.resize(image, (600, 600), interpolation=cv2.INTER_CUBIC)
                
                image = preprocess_input(image)
                image = np.expand_dims(image, axis=0)
                image = datagen.random_transform(image[0])

                # Add Gaussian noise with random standard deviation
                random_noise_std = np.random.uniform(0, augment_params['max_noise_std'])
                image += np.random.normal(0, random_noise_std, image.shape)

                spatial_embedding = feature_extractor.predict(np.expand_dims(image, axis=0))[0]
                features_listofList.append(spatial_embedding)
            
            features_tensor = torch.tensor(features_listofList)
            all_features.append(features_tensor)

        return all_features
    else:
        logging.info("No match found for: " + input_string + "\n")
        return None

def validate_and_pad_data(data):
    if not data:
        return []
    
    max_length = max(item['sign'].shape[0] for item in data)
    expected_shape = data[0]['sign'].shape[1]
    
    for item in data:
        tensor_shape = item['sign'].shape
        if tensor_shape[0] < max_length:
            padding = torch.zeros((max_length - tensor_shape[0], expected_shape))
            item['sign'] = torch.cat((item['sign'], padding), dim=0)
        elif tensor_shape[0] > max_length:
            item['sign'] = item['sign'][:max_length]
        
    return data

def create_pickle(config, frame_dest, checkpoint_path, augment_params, num_augmentations):
    vw_dest = config['vw_dest']
    vo_dest = config['vo_dest']
    checkpoint_name = config['checkpoint_name']

    feature_extractor = initialize_model()

    workbook = load_workbook(os.path.join(os.path.dirname(os.path.abspath(__file__)), vw_dest))
    sheet = workbook.active
    excel_data = []
    for row in sheet.iter_rows(values_only=True):
        excel_data.append(row)

    list_of_inputs, processed_names = load_checkpoint(checkpoint_path, checkpoint_name)
    if list_of_inputs is None:
        list_of_inputs = []

    checkpoint_range = 5
    none_counter = 0
    flag = 0
    reprocessed_entries = []  # List to track reprocessed entries

    for index in range(len(list_of_inputs), len(excel_data), checkpoint_range):
        if flag == 1:
            break
        batch_list_of_inputs = []
        for tmp in excel_data[index:index + checkpoint_range]:
            if tmp[0] in processed_names:
                continue
            augmented_features = get_features(str(tmp[0]), frame_dest, feature_extractor, augment_params, num_augmentations)
            if augmented_features is not None:
                none_counter = 0
                for aug_index, features in enumerate(augmented_features):
                    if len(features) > 0:
                        data_dict = {
                            'name': f"{tmp[0]}_{aug_index + 1}",
                            'signer': tmp[1],
                            'gloss': tmp[2],
                            'text': tmp[3],
                            'sign': features + 1e-8
                        }
                        batch_list_of_inputs.append(data_dict)
                        processed_names.add(f"{tmp[0]}_{aug_index + 1}")
                        if tmp[0] in processed_names:
                            reprocessed_entries.append(tmp[0])
            else:
                none_counter += 1
                if none_counter >= 2 * checkpoint_range - 1:
                    flag = 1
                    break
        if flag == 1:
            break
        
        list_of_inputs.extend(batch_list_of_inputs)

        list_of_inputs = validate_and_pad_data(list_of_inputs)
        
        save_checkpoint(checkpoint_path, checkpoint_name, list_of_inputs)
        torch.cuda.empty_cache()
        list_of_inputs, processed_names = load_checkpoint(checkpoint_path, checkpoint_name)

    if reprocessed_entries:
        logging.info(f"Reprocessed entries: {reprocessed_entries}")

    with open(os.path.join(os.path.dirname(os.path.abspath(__file__)), vo_dest), 'wb') as f:
        with gzip.GzipFile(fileobj=f, mode='wb') as gz:
            pickle.dump(list_of_inputs, gz)

    logging.info(f"Done processing {vw_dest}")

def run_multiple_pickle_creations(configurations, frame_dest, checkpoint_path, augment_params, num_augmentations):
    with ProcessPoolExecutor() as executor:
        futures = []
        for config in configurations:
            logging.info(f"Submitting {config['vw_dest']} -> {config['vo_dest']} with checkpoint {config['checkpoint_name']}")
            futures.append(executor.submit(create_pickle, config, frame_dest, checkpoint_path, augment_params, num_augmentations))
        
        for future in as_completed(futures):
            try:
                future.result()
            except Exception as e:
                logging.error(f"Error occurred: {e}")

if __name__ == "__main__":
    multiprocessing.freeze_support()

    # List of configurations to process
    configurations = [
        {
            'vw_dest': "Dataset/excels/split_file_20.xlsx",
            'vo_dest': "Dataset/Pickles/split_file_complete_20.pickle",
            'checkpoint_name': 'split_file_20.pkl'
        },

        {
            'vw_dest': "Dataset/excels/split_file_1.xlsx",
            'vo_dest': "Dataset/Pickles/split_file_complete_1.pickle",
            'checkpoint_name': 'split_file_1.pkl'
        },
        {
            'vw_dest': "Dataset/excels/split_file_2.xlsx",
            'vo_dest': "Dataset/Pickles/split_file_complete_2.pickle",
            'checkpoint_name': 'split_file_2.pkl'
        },

        {
            'vw_dest': "Dataset/excels/split_file_3.xlsx",
            'vo_dest': "Dataset/Pickles/split_file_complete_3.pickle",
            'checkpoint_name': 'split_file_3.pkl'
        },
        {
            'vw_dest': "Dataset/excels/split_file_4.xlsx",
            'vo_dest': "Dataset/Pickles/split_file_complete_4.pickle",
            'checkpoint_name': 'split_file_4.pkl'
        },

        {
            'vw_dest': "Dataset/excels/split_file_5.xlsx",
            'vo_dest': "Dataset/Pickles/split_file_complete_5.pickle",
            'checkpoint_name': 'split_file_5.pkl'
        },
        {
            'vw_dest': "Dataset/excels/split_file_6.xlsx",
            'vo_dest': "Dataset/Pickles/split_file_complete_6.pickle",
            'checkpoint_name': 'split_file_6.pkl'
        },

        {
            'vw_dest': "Dataset/excels/split_file_7.xlsx",
            'vo_dest': "Dataset/Pickles/split_file_complete_7.pickle",
            'checkpoint_name': 'split_file_7.pkl'
        },
        {
            'vw_dest': "Dataset/excels/split_file_8.xlsx",
            'vo_dest': "Dataset/Pickles/split_file_complete_8.pickle",
            'checkpoint_name': 'split_file_8.pkl'
        },

        {
            'vw_dest': "Dataset/excels/split_file_9.xlsx",
            'vo_dest': "Dataset/Pickles/split_file_complete_9.pickle",
            'checkpoint_name': 'split_file_9.pkl'
        },

        {
            'vw_dest': "Dataset/excels/split_file_10.xlsx",
            'vo_dest': "Dataset/Pickles/split_file_complete_10.pickle",
            'checkpoint_name': 'split_file_10.pkl'
        },

        {
            'vw_dest': "Dataset/excels/split_file_11.xlsx",
            'vo_dest': "Dataset/Pickles/split_file_complete_11.pickle",
            'checkpoint_name': 'split_file_11.pkl'
        },
        {
            'vw_dest': "Dataset/excels/split_file_12.xlsx",
            'vo_dest': "Dataset/Pickles/split_file_complete_12.pickle",
            'checkpoint_name': 'split_file_12.pkl'
        },

        {
            'vw_dest': "Dataset/excels/split_file_13.xlsx",
            'vo_dest': "Dataset/Pickles/split_file_complete_13.pickle",
            'checkpoint_name': 'split_file_13.pkl'
        },
        {
            'vw_dest': "Dataset/excels/split_file_14.xlsx",
            'vo_dest': "Dataset/Pickles/split_file_complete_14.pickle",
            'checkpoint_name': 'split_file_14.pkl'
        },

        {
            'vw_dest': "Dataset/excels/split_file_15.xlsx",
            'vo_dest': "Dataset/Pickles/split_file_complete_15.pickle",
            'checkpoint_name': 'split_file_15.pkl'
        },
        {
            'vw_dest': "Dataset/excels/split_file_16.xlsx",
            'vo_dest': "Dataset/Pickles/split_file_complete_16.pickle",
            'checkpoint_name': 'split_file_16.pkl'
        },

        {
            'vw_dest': "Dataset/excels/split_file_17.xlsx",
            'vo_dest': "Dataset/Pickles/split_file_complete_17.pickle",
            'checkpoint_name': 'split_file_17.pkl'
        },
        {
            'vw_dest': "Dataset/excels/split_file_18.xlsx",
            'vo_dest': "Dataset/Pickles/split_file_complete_18.pickle",
            'checkpoint_name': 'split_file_18.pkl'
        },
        {
            'vw_dest': "Dataset/excels/split_file_19.xlsx",
            'vo_dest': "Dataset/Pickles/split_file_complete_19.pickle",
            'checkpoint_name': 'split_file_19.pkl'
        }
    ]

    vf_dest = "Dataset/Final folder for frames"
    store_to_path = 'C:\\Users\\Admin\\Rahul\\islt_multi_modality_phoenix_dataset\\Pre-Processing\\Pickle maker\\Dataset\\Checkpoint\\'

    # Variables for augmentation
    augment_params = {
        'brightness': 0.15,  # 15% brightness change
        'rotation': 20,  # 20 degrees rotation
        'width_shift': 0.2,  # 20% width shift
        'height_shift': 0.2,  # 20% height shift
        'zoom': 0.15,  # 15% zoom
        'shear': 0.15,  # 15% shear
        'max_noise_std': 0.1  # Maximum standard deviation of Gaussian noise
    }
    num_augmentations = 25  # Number of augmented copies

    run_multiple_pickle_creations(configurations, vf_dest, store_to_path, augment_params, num_augmentations)

    logging.info("Done creating pickle files for all configurations.")